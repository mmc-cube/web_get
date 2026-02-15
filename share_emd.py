"""闲鱼接单文案生成 → Telegram 推送（纯文字版）"""

import os
import random
import requests
from datetime import datetime

import openpyxl


# =============================
# 配置（优先读环境变量）
# =============================

XLSX_PATH = os.getenv("SHARE_XLSX_PATH", "data/share_modules.xlsx")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "").strip()

QWEN_API_KEY = os.getenv("QWEN_API_KEY", os.getenv("DASHSCOPE_API_KEY", "")).strip()
QWEN_BASE_URL = os.getenv("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1").rstrip("/")
QWEN_MODEL = os.getenv("QWEN_MODEL", "qwen-plus")


# =============================
# Excel 读取（用 openpyxl，不依赖 pandas）
# =============================

def load_sheets(path: str) -> dict:
    """读取 xlsx，返回 {sheet_name: [rows]}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip().lower() if h else f"col{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            d = {headers[i]: (str(v).strip() if v else "") for i, v in enumerate(row) if i < len(headers)}
            if any(d.values()):
                data.append(d)
        result[name.lower()] = data
    wb.close()
    return result


def pick_one(rows: list[dict], col: str) -> str:
    return random.choice(rows)[col]


def pick_many(rows: list[dict], col: str, k: int) -> list[str]:
    k = min(k, len(rows))
    return [r[col] for r in random.sample(rows, k) if r.get(col)]


# =============================
# 文案组装
# =============================

def build_raw_text(sheets: dict) -> str:
    modules = sheets.get("modules", [])
    projects = sheets.get("projects", [])

    if not modules:
        raise ValueError("modules sheet 为空")

    def pick_type(t: str) -> str:
        sub = [r for r in modules if r.get("type", "").lower() == t]
        if not sub:
            raise ValueError(f"modules 里找不到 type={t}")
        return pick_one(sub, "text")

    hook = pick_type("hook")
    ability = pick_type("ability")
    service = pick_type("service")
    deliver1 = pick_type("deliver")
    deliver2 = pick_type("deliver")
    trust = pick_type("trust")
    cta = pick_type("cta")

    # 防止 deliver 重复
    deliver_pool = [r for r in modules if r.get("type", "").lower() == "deliver"]
    if deliver2 == deliver1 and len(deliver_pool) > 1:
        for _ in range(5):
            d = pick_type("deliver")
            if d != deliver1:
                deliver2 = d
                break

    proj_col = list(projects[0].keys())[0] if projects else "col0"
    proj_list = pick_many(projects, proj_col, 6) if projects else []

    raw = (
        f"{hook}\n\n"
        f"{ability}\n"
        f"{service}\n\n"
    )
    if proj_list:
        raw += f"可做方向（随机举例）：{' / '.join(proj_list)}\n\n"
    raw += (
        f"交付：\n"
        f"- {deliver1}\n"
        f"- {deliver2}\n\n"
        f"{trust}\n\n"
        f"{cta}"
    )
    return raw


# =============================
# Qwen 润色
# =============================

def qwen_polish(text: str) -> str:
    if not QWEN_API_KEY:
        return text

    url = f"{QWEN_BASE_URL}/chat/completions"
    headers = {
        "Authorization": f"Bearer {QWEN_API_KEY}",
        "Content-Type": "application/json",
    }

    user_prompt = (
        "你要做的是"轻度润色"，不是重写。\n\n"
        "规则（必须遵守）：\n"
        "1) 只允许调整语序、合并/拆分句子、删掉重复表达；不要新增能力、不要新增项目、不要编造数据或案例。\n"
        "2) 保留原文信息点：能力范围、可做方向、交付、背书、引导私信。\n"
        "3) 输出像真实工程师发闲鱼，不要像广告；不要用夸张词：如"轻松搞定/从0到1/完美/秒出/全网最低"等。\n"
        "4) 不要写强营销口号（如"限时/速来/赶紧下单"）。可以保留一句自然的私信引导。\n"
        "5) 不要使用小标题（如【交付内容】），用自然段即可。\n"
        "6) 字数控制在 220~320 字。\n"
        "7) 重点 我们行业的主要获客关键词就是单片机设计 物联网开发 stm32 开发 esp32单片机开发 "
        "禁止出现关于课程设计 毕业设计 大学生毕业设计 大学生课程设计 等等这种学术代做的文字\n\n"
        "把下面文本润色后输出（只输出润色后的最终文案）：\n\n"
        f"<<<\n{text}\n>>>"
    )

    payload = {
        "model": QWEN_MODEL,
        "messages": [
            {"role": "system", "content": "你是资深嵌入式工程师文案助手，风格自然、克制、可信。"},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.7
    }

    r = requests.post(url, headers=headers, json=payload, timeout=60)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip()


# =============================
# Telegram 发送
# =============================

def telegram_send_text(msg: str):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print(f"[Telegram 未配置] {msg[:80]}...")
        return

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TELEGRAM_CHAT_ID, "text": msg}
    r = requests.post(url, json=payload, timeout=30)
    r.raise_for_status()


# =============================
# 主流程：生成 N 条 + 汇报
# =============================

def run_batch(count: int = 4):
    """生成 count 条文案，带编号发送，最后发日期汇报"""
    sheets = load_sheets(XLSX_PATH)

    success = 0
    for i in range(1, count + 1):
        try:
            raw = build_raw_text(sheets)
            final = qwen_polish(raw) if QWEN_API_KEY else raw

            # 加编号前缀
            numbered = f"【{i}/{count}】\n\n{final}"
            telegram_send_text(numbered)
            success += 1
            print(f"[{i}/{count}] 已发送")
        except Exception as e:
            print(f"[{i}/{count}] 失败: {e}")

    # 发送日期汇报
    today = datetime.now().strftime("%Y-%m-%d %H:%M")
    report = f"--- 今日文案推送完毕 ---\n日期: {today}\n成功: {success}/{count} 条"
    telegram_send_text(report)
    print(report)


if __name__ == "__main__":
    import sys
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 4
    run_batch(n)
